from openpyxl import load_workbook

class ExcelFunctions:

    def _init_(self,file_name,sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
   
    # method to get the column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column
   
    # method to read data from the excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value


    # method to write the data into the excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)




