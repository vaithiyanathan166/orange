import pytest
import openpyxl
from pages.login_page import LoginPage

URL = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"

@pytest.mark.usefixtures("driver")
class TestLoginDDTF:
    def test_login_using_excel(self, driver):
        login_page = LoginPage(driver)
        login_page.open_url(URL)
        
        # Load data from Excel
        workbook = openpyxl.load_workbook("test_data/login_data.xlsx")
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):
            username = sheet.cell(row, 1).value
            password = sheet.cell(row, 2).value
            expected_result = sheet.cell(row, 3).value

            login_page.login(username, password)
            actual_result = "Pass" if login_page.is_login_successful() else "Fail"
            assert actual_result == expected_result